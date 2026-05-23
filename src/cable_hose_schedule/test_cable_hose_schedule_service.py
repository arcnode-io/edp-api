"""CableHoseScheduleService unit tests — derives cable + hose lists from DTM."""

import json as _json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.cable_hose_schedule.cable_hose_schedule_models import CableHoseSchedule
from src.cable_hose_schedule.cable_hose_schedule_service import (
    CableHoseScheduleService,
    serialize_cable_hose_schedule_xlsx,
)
from src.drawing.conftest import make_device, make_dtm
from src.shared.schemas.dtm import Dtm
from src.shared.schemas.measurement import Measurement, Publisher
from src.shared.schemas.template import ContainsEntry, DeviceTemplate, TemplateKind
from src.shared.schemas.template_protocols import (
    Dnp3Binding,
    ModbusBinding,
    RedfishBinding,
    SnmpBinding,
)


def _tpl(
    slug: str,
    binding: ModbusBinding | Dnp3Binding | RedfishBinding | SnmpBinding,
) -> DeviceTemplate:
    return DeviceTemplate(
        template=slug,
        kind=TemplateKind.LEAF,
        equipment_id=f"EXT-{slug.upper()}-001",
        vendor="V",
        model="M",
        description=slug,
        measurements={
            "power": Measurement(unit="watts", type="float", binding=binding)
        },
    )


@pytest.fixture
def three_protocol_dtm() -> Dtm:
    """3 devices on 3 different protocols — each yields one comms cable."""
    return make_dtm(
        devices={
            "switchgear_1": make_device("switchgear_1", template="switchgear"),
            "relay_1": make_device("relay_1", template="protective_relay"),
            "gpu_node_1": make_device("gpu_node_1", template="gpu_node"),
        },
        templates={
            "switchgear": _tpl(
                "switchgear",
                ModbusBinding(protocol="modbus_tcp", function_code=3, address=0),
            ),
            "protective_relay": _tpl(
                "protective_relay",
                Dnp3Binding(
                    protocol="dnp3_tcp", point_index=0, point_type="analog_input"
                ),
            ),
            "gpu_node": _tpl(
                "gpu_node", RedfishBinding(protocol="redfish", uri="/redfish/v1")
            ),
        },
    )


def test_generate_returns_cable_hose_schedule(three_protocol_dtm: Dtm) -> None:
    """generate() returns a CableHoseSchedule Pydantic object."""
    # Act
    actual = CableHoseScheduleService().generate(three_protocol_dtm)

    # Assert
    assert isinstance(actual, CableHoseSchedule)


def test_one_cable_per_bound_device(three_protocol_dtm: Dtm) -> None:
    """Every device with a Binding gets one comms cable row."""
    # Act
    schedule = CableHoseScheduleService().generate(three_protocol_dtm)

    # Assert
    assert len(schedule.cables) == 3
    device_ids = {c.from_device for c in schedule.cables}
    assert device_ids == set(three_protocol_dtm.devices)


def test_cable_carries_protocol_specific_type(three_protocol_dtm: Dtm) -> None:
    """Cable `cable_type` reflects the protocol (e.g. Modbus TCP → Cat6)."""
    # Act
    schedule = CableHoseScheduleService().generate(three_protocol_dtm)
    by_device = {c.from_device: c for c in schedule.cables}

    # Assert
    assert "Cat6" in by_device["switchgear_1"].cable_type
    assert "Cat6" in by_device["relay_1"].cable_type
    assert "Cat6" in by_device["gpu_node_1"].cable_type


def test_json_serialization_round_trips(three_protocol_dtm: Dtm) -> None:
    """`model_dump_json` produces parseable JSON with cables[] + metadata."""
    # Act
    schedule = CableHoseScheduleService().generate(three_protocol_dtm)
    body = _json.loads(schedule.model_dump_json())

    # Assert
    assert body["deployment_uuid"] == str(three_protocol_dtm.deployment_uuid)
    assert "generated_at" in body
    assert len(body["cables"]) == 3
    assert body["hoses"] == []


def test_xlsx_serializer_opens_with_cables_sheet(three_protocol_dtm: Dtm) -> None:
    """xlsx round-trips through openpyxl with a 'Cables' sheet of the right size."""
    # Act
    schedule = CableHoseScheduleService().generate(three_protocol_dtm)
    xlsx = serialize_cable_hose_schedule_xlsx(schedule)
    wb = load_workbook(BytesIO(xlsx))

    # Assert
    assert "Cables" in wb.sheetnames
    cables_ws = wb["Cables"]
    # Header row + one row per cable.
    assert cables_ws.max_row == 1 + 3


def test_cables_terminate_at_local_switch_when_present_in_dtm() -> None:
    """When DTM has a network_switch device, cables terminate there — not at gateway.

    Honest physical topology: comms cables physically run from devices to the
    rack's local network switch. The gateway (cloud or on-prem) sits behind
    that switch and is not a cable endpoint.
    """
    # Arrange
    dtm = make_dtm(
        devices={
            "switch_top": make_device("switch_top", template="network_switch"),
            "switchgear_1": make_device("switchgear_1", template="switchgear"),
            "relay_1": make_device("relay_1", template="protective_relay"),
        },
        templates={
            "network_switch": _tpl(
                "network_switch",
                SnmpBinding(protocol="snmp", oid="1.3.6.1.2.1.1.5.0"),
            ),
            "switchgear": _tpl(
                "switchgear",
                ModbusBinding(protocol="modbus_tcp", function_code=3, address=0),
            ),
            "protective_relay": _tpl(
                "protective_relay",
                Dnp3Binding(
                    protocol="dnp3_tcp", point_index=0, point_type="analog_input"
                ),
            ),
        },
    )

    # Act
    schedule = CableHoseScheduleService().generate(dtm)

    # Assert — every cable terminates at the switch, not "industrial_gateway".
    # The switch itself is not its own cable row.
    cable_to_devices = {c.to_device for c in schedule.cables}
    cable_from_devices = {c.from_device for c in schedule.cables}
    assert cable_to_devices == {"switch_top"}
    assert "switch_top" not in cable_from_devices
    # 2 cables (one per non-switch device).
    assert len(schedule.cables) == 2


def test_cables_fall_back_to_tbd_label_when_no_switch_in_dtm(
    three_protocol_dtm: Dtm,
) -> None:
    """No network_switch in DTM → to_device labelled as TBD so the gap is honest."""
    # Act
    schedule = CableHoseScheduleService().generate(three_protocol_dtm)

    # Assert — every cable's to_device explicitly flags the missing-switch gap
    for cable in schedule.cables:
        assert "TBD" in cable.to_device


def _compute_module_dtm_with_pdu_and_cdu() -> Dtm:
    """A realistic compute_module instance: 1 cdu + 2 gpu_nodes + 1 pdu + 1 switch.

    Mirrors the real `device_templates/module/compute_module.yaml` shape:
    PDU is the power source; cdu, gpu_node, network_switch all carry
    `power_from: [pdu]` on their compute_module contains entry.
    """
    # Build templates that mirror the real ones (contains[].power_from drives
    # power-cable derivation; cdu+gpu_node co-residence drives hoses).
    pdu_tpl = _tpl(
        "pdu", SnmpBinding(protocol="snmp", oid="1.3.6.1.4.1.1718.4.1.3.3.1.7")
    )
    cdu_tpl = _tpl("cdu", RedfishBinding(protocol="redfish", uri="/redfish/v1"))
    gpu_node_tpl = _tpl(
        "gpu_node", RedfishBinding(protocol="redfish", uri="/redfish/v1/Systems/1")
    )
    switch_tpl = _tpl(
        "network_switch",
        SnmpBinding(protocol="snmp", oid="1.3.6.1.2.1.1.5.0"),
    )
    compute_module_tpl = DeviceTemplate(
        template="compute_module",
        kind=TemplateKind.MODULE,
        description="compute_module",
        contains=[
            ContainsEntry(template="pdu", qty="scalable"),
            ContainsEntry(template="gpu_node", qty="scalable", power_from=["pdu"]),
            ContainsEntry(template="cdu", qty=1, power_from=["pdu"]),
            ContainsEntry(
                template="network_switch", qty="scalable", power_from=["pdu"]
            ),
        ],
        measurements={
            "total_power": Measurement(
                unit="watts", type="float", publisher=Publisher.LINE_CONTROLLER
            )
        },
    )

    return make_dtm(
        devices={
            "compute_module_1": make_device(
                "compute_module_1", template="compute_module"
            ).model_copy(update={"connection": None}),
            "pdu_1": make_device("pdu_1", template="pdu").model_copy(
                update={"parent": "compute_module_1"}
            ),
            "cdu_1": make_device("cdu_1", template="cdu").model_copy(
                update={"parent": "compute_module_1"}
            ),
            "gpu_node_1": make_device("gpu_node_1", template="gpu_node").model_copy(
                update={"parent": "compute_module_1"}
            ),
            "gpu_node_2": make_device("gpu_node_2", template="gpu_node").model_copy(
                update={"parent": "compute_module_1"}
            ),
            "switch_top": make_device(
                "switch_top", template="network_switch"
            ).model_copy(update={"parent": "compute_module_1"}),
        },
        templates={
            "compute_module": compute_module_tpl,
            "pdu": pdu_tpl,
            "cdu": cdu_tpl,
            "gpu_node": gpu_node_tpl,
            "network_switch": switch_tpl,
        },
    )


def test_power_cable_per_powered_device() -> None:
    """Every device whose contains[] entry has `power_from` gets a power cable row.

    From-device = a PDU sibling; to-device = the powered device. The PDU
    itself doesn't get a power cable (its upstream feed is out-of-scope v1).
    """
    # Arrange
    dtm = _compute_module_dtm_with_pdu_and_cdu()

    # Act
    schedule = CableHoseScheduleService().generate(dtm)

    # Assert — power cables exist for cdu, gpu_node_1, gpu_node_2, switch_top.
    # NOT for pdu_1 (it's the source) or compute_module_1 (no connection).
    power_cables = [c for c in schedule.cables if c.service.startswith("Power")]
    powered = {c.to_device for c in power_cables}
    assert powered == {"cdu_1", "gpu_node_1", "gpu_node_2", "switch_top"}
    # All power cables originate at pdu_1.
    assert {c.from_device for c in power_cables} == {"pdu_1"}


def test_cdu_primary_hoses_are_by_others_facility_side() -> None:
    """CDU primary supply + return are BY OTHERS (facility-side coolant)."""
    # Arrange
    dtm = _compute_module_dtm_with_pdu_and_cdu()

    # Act
    schedule = CableHoseScheduleService().generate(dtm)

    # Assert
    primary_hoses = [h for h in schedule.hoses if "Primary" in h.service]
    assert len(primary_hoses) == 2  # supply + return
    for hose in primary_hoses:
        assert "BY OTHERS" in hose.notes or "BY OTHERS" in hose.to_device


def test_dlc_plate_hoses_per_gpu_node() -> None:
    """Each gpu_node co-resident with a CDU gets a secondary supply + return hose."""
    # Arrange
    dtm = _compute_module_dtm_with_pdu_and_cdu()

    # Act
    schedule = CableHoseScheduleService().generate(dtm)

    # Assert — 2 hoses per gpu_node (supply + return). 2 gpu_nodes → 4 hoses.
    secondary_hoses = [h for h in schedule.hoses if "Secondary" in h.service]
    assert len(secondary_hoses) == 4
    # Each plate appears as either from_device (return path) or to_device
    # (supply path); intersect with non-CDU devices to get the plate set.
    plate_devices = {
        d for h in secondary_hoses for d in (h.from_device, h.to_device)
    } - {"cdu_1"}
    assert plate_devices == {"gpu_node_1", "gpu_node_2"}


def test_xlsx_has_hoses_sheet_even_when_empty(three_protocol_dtm: Dtm) -> None:
    """`Hoses` sheet exists with the header row even when DTM has no liquid buses.

    v1 doesn't model liquid buses; hoses become derivable when the bus
    schema grows a `liquid` type. The empty sheet is the explicit signal
    to a reviewer that hose derivation is on the roadmap.
    """
    # Act
    schedule = CableHoseScheduleService().generate(three_protocol_dtm)
    xlsx = serialize_cable_hose_schedule_xlsx(schedule)
    wb = load_workbook(BytesIO(xlsx))

    # Assert
    assert "Hoses" in wb.sheetnames
    hoses_ws = wb["Hoses"]
    assert hoses_ws.max_row == 1  # header row only
