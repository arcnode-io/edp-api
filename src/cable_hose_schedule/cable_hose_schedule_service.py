"""CableHoseScheduleService — derives cable + hose schedule from the DTM.

Produces JSON (canonical) + XLSX (reviewer-friendly) outputs at the
reserved cable_hose_schedule URLs.

Three derivations:

1. **Comms cables** — one per device with a measurement binding. From
   `Device.connection` (host:port + optional unit_id) to the local
   `network_switch` device (port GE0/0/N). The gateway (cloud or on-prem)
   sits behind the switch and is never a cable endpoint.

2. **Power cables** — driven by `ContainsEntry.power_from` on the module
   template. For every powered child, emits one row from a PDU sibling
   (primary feed) to the consumer device. Redundant 2N feeds are
   reserved for v2 when PDU instances carry A/B-side designators.

3. **Hoses** — every CDU device in a parent module emits a primary
   supply + return pair as BY-OTHERS (facility loop). Every gpu_node
   co-resident with that CDU also emits a secondary supply + return
   pair to its DLC plate. Mirrors the P&ID's coolant routing — same
   underlying data.

Fallbacks:
- No `network_switch` device → comms cable `to_device` is `TBD-LOCAL-SWITCH`.
- No PDU sibling → power cable not emitted (honest gap; surfaces as
  missing rows rather than fake from-device labels).
"""

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.cable_hose_schedule.cable_hose_schedule_models import (
    CableEntry,
    CableHoseSchedule,
    HoseEntry,
)
from src.shared.schemas.dtm import Device, Dtm
from src.shared.schemas.template import DeviceTemplate

# Protocol → cable type mapping. All IP-based protocols share Cat6 STP at v1
# (could split out fiber per spec later). CANopen is canbus-twisted-pair.
_PROTOCOL_CABLE_TYPE: dict[str, tuple[str, str]] = {
    # protocol -> (service label, cable type)
    "modbus_tcp": ("Comms - Modbus TCP", "Cat6 STP"),
    "dnp3_tcp": ("Comms - DNP3 TCP", "Cat6 STP"),
    "snmp": ("Comms - SNMP", "Cat6 STP"),
    "redfish": ("Comms - Redfish", "Cat6 STP"),
    "canopen_gw": ("Comms - CANopen", "CAN bus twisted-pair"),
}

# Fallback when no `network_switch` device exists in the DTM. Honest signal
# that the cable termination is unresolved — better than naming a switch
# we don't actually know exists.
_TBD_SWITCH_DEVICE_ID: str = "TBD-LOCAL-SWITCH"
_TBD_SWITCH_PORT: str = "TBD"

# Template slug that identifies the rack's local network switch. v1 has
# one switch template; if a deployment grows multiple switch templates,
# this will need to be a set lookup.
_SWITCH_TEMPLATE_SLUG: str = "network_switch"

_CABLE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("tag", "Tag"),
    ("service", "Service"),
    ("from_device", "From Device"),
    ("from_port", "From Port"),
    ("to_device", "To Device"),
    ("to_port", "To Port"),
    ("cable_type", "Cable Type"),
    ("length_estimate_m", "Length (m)"),
    ("notes", "Notes"),
)

_HOSE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("tag", "Tag"),
    ("service", "Service"),
    ("from_device", "From Device"),
    ("from_port", "From Port"),
    ("to_device", "To Device"),
    ("to_port", "To Port"),
    ("hose_type", "Hose Type"),
    ("length_estimate_m", "Length (m)"),
    ("notes", "Notes"),
)


class CableHoseScheduleService:
    """Walks the DTM, emits one comms cable per bound device.

    Returns a `CableHoseSchedule` Pydantic object — the caller picks the
    serialization format (mirrors `BomGeneratorService` which returns
    `Bom` + has module-level `serialize_bom_xlsx`). Pipeline dispatch
    calls `model_dump_json` for the json URL and
    `serialize_cable_hose_schedule_xlsx` for the xlsx URL.
    """

    def generate(self, dtm: Dtm) -> CableHoseSchedule:
        """Build the schedule. Caller serializes to JSON / XLSX."""
        return self._build_schedule(dtm)

    def _build_schedule(self, dtm: Dtm) -> CableHoseSchedule:
        switch_device_id = _find_local_switch(dtm)
        cables = _comms_cables(dtm, switch_device_id) + _power_cables(dtm)
        # Renumber tags sequentially across both sources for a clean output.
        cables = [
            cable.model_copy(update={"tag": f"CBL-{i + 1:04d}"})
            for i, cable in enumerate(cables)
        ]
        hoses = _hoses(dtm)
        hoses = [
            hose.model_copy(update={"tag": f"HOS-{i + 1:04d}"})
            for i, hose in enumerate(hoses)
        ]
        return CableHoseSchedule(
            deployment_uuid=dtm.deployment_uuid,
            generated_at=datetime.now(UTC),
            cables=cables,
            hoses=hoses,
        )


def _comms_cables(dtm: Dtm, switch_device_id: str | None) -> list[CableEntry]:
    """One comms cable per bound device, terminating at the local switch."""
    cables: list[CableEntry] = []
    tag_num = 0
    for device_id in sorted(dtm.devices):
        device = dtm.devices[device_id]
        if device.connection is None or device_id == switch_device_id:
            continue
        template = dtm.templates_used.get(device.template)
        tag_num += 1
        cable = _comms_cable_for_device(
            template,
            device_id,
            device,
            tag_num=tag_num,
            switch_device_id=switch_device_id,
            switch_port_num=tag_num,
        )
        if cable is not None:
            cables.append(cable)
        else:
            tag_num -= 1
    return cables


def _power_cables(dtm: Dtm) -> list[CableEntry]:
    """One power cable per device whose contains[] entry has `power_from`.

    Walks every module device in the DTM, reads its template's contains[]
    entries, and for each entry with `power_from` finds the first sibling
    device whose template matches a power_from slug. Emits a power-cable
    row from that sibling to the powered child device. Devices with no
    suitable PDU sibling are skipped (honest gap; surfaces as missing
    rows rather than fake from-device labels).
    """
    cables: list[CableEntry] = []
    tag_num = 0
    for module_id in sorted(dtm.devices):
        module = dtm.devices[module_id]
        module_template = dtm.templates_used.get(module.template)
        if module_template is None or not module_template.contains:
            continue
        # Map contains[].template -> power_from list for quick lookup.
        power_map = {
            entry.template: entry.power_from for entry in module_template.contains
        }
        # Children of this module instance, keyed by template slug.
        children_by_template = _children_by_template(dtm, module_id)
        for child_device in _powered_children(power_map, children_by_template):
            powered_template = child_device.template
            sources = power_map[powered_template]
            source_device = _first_source_sibling(sources, children_by_template)
            if source_device is None:
                continue
            tag_num += 1
            cables.append(
                CableEntry(
                    tag=f"PWR-{tag_num:04d}",  # placeholder; reassigned in _build_schedule
                    service="Power - 240V AC (single-phase from PDU outlet)",
                    from_device=source_device.device_id,
                    from_port="C13 outlet (sequential)",
                    to_device=child_device.device_id,
                    to_port="PWR (PSU input)",
                    cable_type="C13-C14 cordset, AWG 14 SJT",
                    length_estimate_m=None,
                    notes="Primary feed only; redundant B-side feed reserved for v2.",
                )
            )
    return cables


def _hoses(dtm: Dtm) -> list[HoseEntry]:
    """Per CDU: primary supply+return BY OTHERS + secondary supply+return per gpu_node.

    Walks every module device in the DTM, finds the CDU child (if any) and
    the gpu_node children. Emits:
    - 2 hoses per CDU (primary supply + return), BY OTHERS facility side
    - 2 hoses per (cdu, gpu_node) pair (secondary supply + return)
    """
    hoses: list[HoseEntry] = []
    tag = 0
    for module_id in sorted(dtm.devices):
        children_by_template = _children_by_template(dtm, module_id)
        cdus = children_by_template.get("cdu", [])
        gpu_nodes = children_by_template.get("gpu_node", [])
        for cdu in cdus:
            # Primary supply + return — BY OTHERS (facility loop).
            tag += 1
            hoses.append(
                HoseEntry(
                    tag=f"HOS-{tag:04d}",
                    service="Coolant Primary Supply - Facility Loop",
                    from_device="FACILITY (BY OTHERS)",
                    from_port="TBD",
                    to_device=cdu.device_id,
                    to_port="PRI_SUPPLY",
                    hose_type="EPDM 1in PG/W rated",
                    length_estimate_m=None,
                    notes="BY OTHERS - facility cooling loop provided by customer.",
                )
            )
            tag += 1
            hoses.append(
                HoseEntry(
                    tag=f"HOS-{tag:04d}",
                    service="Coolant Primary Return - Facility Loop",
                    from_device=cdu.device_id,
                    from_port="PRI_RETURN",
                    to_device="FACILITY (BY OTHERS)",
                    to_port="TBD",
                    hose_type="EPDM 1in PG/W rated",
                    length_estimate_m=None,
                    notes="BY OTHERS - facility cooling loop provided by customer.",
                )
            )
            # Secondary supply + return per gpu_node DLC plate.
            for gpu_node in gpu_nodes:
                tag += 1
                hoses.append(
                    HoseEntry(
                        tag=f"HOS-{tag:04d}",
                        service="Coolant Secondary Supply - Rack DLC",
                        from_device=cdu.device_id,
                        from_port="SEC_SUPPLY (via manifold)",
                        to_device=gpu_node.device_id,
                        to_port="GPU_LIQ_SUPPLY",
                        hose_type="EPDM 0.5in DEI water rated",
                        length_estimate_m=None,
                        notes="Via rack-rear blind-mate manifold.",
                    )
                )
                tag += 1
                hoses.append(
                    HoseEntry(
                        tag=f"HOS-{tag:04d}",
                        service="Coolant Secondary Return - Rack DLC",
                        from_device=gpu_node.device_id,
                        from_port="GPU_LIQ_RETURN",
                        to_device=cdu.device_id,
                        to_port="SEC_RETURN (via manifold)",
                        hose_type="EPDM 0.5in DEI water rated",
                        length_estimate_m=None,
                        notes="Via rack-rear blind-mate manifold.",
                    )
                )
    return hoses


def _children_by_template(dtm: Dtm, parent_device_id: str) -> dict[str, list[Device]]:
    """Group a parent's direct children by template slug."""
    result: dict[str, list[Device]] = {}
    for did in sorted(dtm.devices):
        device = dtm.devices[did]
        if device.parent != parent_device_id:
            continue
        result.setdefault(device.template, []).append(device)
    return result


def _powered_children(
    power_map: dict[str, list[str]],
    children_by_template: dict[str, list[Device]],
) -> list[Device]:
    """Devices whose contains entry declares one or more `power_from` sources."""
    powered: list[Device] = []
    for template_slug, sources in power_map.items():
        if not sources:
            continue
        powered.extend(children_by_template.get(template_slug, []))
    return powered


def _first_source_sibling(
    source_template_slugs: list[str],
    children_by_template: dict[str, list[Device]],
) -> Device | None:
    """First device under the same module whose template is in source_template_slugs."""
    for slug in source_template_slugs:
        candidates = children_by_template.get(slug, [])
        if candidates:
            return candidates[0]
    return None


def _comms_cable_for_device(
    template: DeviceTemplate | None,
    device_id: str,
    device: Device,
    *,
    tag_num: int,
    switch_device_id: str | None,
    switch_port_num: int,
) -> CableEntry | None:
    """Build one comms cable from a device's template binding + connection.

    Cable terminates at `switch_device_id` if the DTM includes a
    network_switch device; otherwise at the explicit `TBD-LOCAL-SWITCH`
    placeholder. `switch_port_num` sequentially numbers the switch's
    front-panel ports (GE0/0/N) — assignment is deterministic for a
    given DTM but doesn't reflect any real cable-routing layout.
    """
    if template is None or not template.measurements:
        return None
    if device.connection is None:
        return None
    protocol: str | None = None
    for measurement in template.measurements.values():
        if measurement.binding is not None:
            protocol = getattr(measurement.binding, "protocol", None)
            if protocol in _PROTOCOL_CABLE_TYPE:
                break
    if protocol not in _PROTOCOL_CABLE_TYPE:
        return None
    service, cable_type = _PROTOCOL_CABLE_TYPE[protocol]
    conn = device.connection
    unit_suffix = f" (unit_id={conn.unit_id})" if conn.unit_id else ""
    to_device = switch_device_id or _TBD_SWITCH_DEVICE_ID
    to_port = f"GE0/0/{switch_port_num}" if switch_device_id else _TBD_SWITCH_PORT
    return CableEntry(
        tag=f"CBL-{tag_num:04d}",
        service=service,
        from_device=device_id,
        from_port=f"{conn.host}:{conn.port}{unit_suffix}",
        to_device=to_device,
        to_port=to_port,
        cable_type=cable_type,
        length_estimate_m=None,  # field-measured per installation
        notes="",
    )


def _find_local_switch(dtm: Dtm) -> str | None:
    """First device whose template is `network_switch`. None if no switch in DTM.

    Deterministic across re-runs (sorted by device_id). v1 expects 0..1
    switches per deployment; multi-switch deployments would need a smarter
    "which device cables to which switch" policy than first-switch-wins.
    """
    for device_id in sorted(dtm.devices):
        if dtm.devices[device_id].template == _SWITCH_TEMPLATE_SLUG:
            return device_id
    return None


def serialize_cable_hose_schedule_xlsx(schedule: CableHoseSchedule) -> bytes:
    """Two sheets — Cables (one row per cable) + Hoses (header row only at v1)."""
    wb = Workbook()
    # Default sheet becomes "Cables".
    cables_ws = wb.active
    cables_ws.title = "Cables"
    _write_table(cables_ws, _CABLE_COLUMNS, schedule.cables)

    hoses_ws = wb.create_sheet(title="Hoses")
    _write_table(hoses_ws, _HOSE_COLUMNS, schedule.hoses)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_table(
    ws: Worksheet,
    columns: tuple[tuple[str, str], ...],
    rows: list[CableEntry] | list[HoseEntry],
) -> None:
    """Write a header row + one data row per Pydantic entry."""
    bold = Font(bold=True)
    for col_idx, (_field, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = bold
    for row_idx, entry in enumerate(rows, start=2):
        for col_idx, (field, _label) in enumerate(columns, start=1):
            value = getattr(entry, field)
            # Reason: openpyxl can't write None — store as blank string.
            ws.cell(
                row=row_idx, column=col_idx, value=value if value is not None else ""
            )
