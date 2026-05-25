"""BOM generator service.

Per Q9-C: fetches manifest, resolves profile→assets, fetches each
referenced spec.yaml + bom.yaml, transforms into bom.json per Q17-A.

Entry: BomGeneratorService.generate(deployment_id, profile, manifest_url,
container_counts) → Bom (uploaded to S3 by caller).
"""

import json
import logging
from datetime import UTC, datetime
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font

from src.bom_generator.bom_models import (
    Bom,
    BomLineItem,
    ProcurementPath,
)
from src.bom_generator.manifest_client import ManifestClient
from src.bom_generator.manifest_models import Manifest, ProfileAssemblies

logger = logging.getLogger(__name__)


def _spec_to_catalog_line(spec: dict, qty: int) -> BomLineItem:
    """Map spec.yaml fields → catalog BOM line item."""
    return BomLineItem(
        part_number=spec.get("model_number", spec["equipment_id"]),
        vendor=spec.get("vendor", "TBD"),
        description=spec.get("description", ""),
        qty=qty,
        procurement_path=ProcurementPath.CATALOG,
        datasheet_url=spec.get("datasheet_url"),
        lead_time_weeks=spec.get("lead_time_weeks"),
        unit_cost_usd=spec.get("unit_cost_usd"),
        fab_tier=spec.get("fab_tier"),
        # Track-A enrichment — install_video_url is a new spec field;
        # NDAA + TAA are DERIVED from existing schema (restricted_entities
        # + fab_tier) to avoid duplicate-source drift.
        install_video_url=spec.get("install_video_url"),
        ndaa_compliant=_derive_ndaa(spec),
        taa_compliant=_derive_taa(spec),
    )


def _derive_ndaa(spec: dict) -> bool:
    """True iff spec.restricted_entities does NOT include `NDAA_889`.

    A spec without restricted_entities (or with an empty list) is by
    convention NDAA-compliant — that's the "checked clean" state per
    the equipment_spec_schema.md doc.
    """
    restricted = spec.get("restricted_entities") or []
    return "NDAA_889" not in restricted


def _derive_taa(spec: dict) -> bool:
    """True iff spec.fab_tier is federal_civilian or dod_eligible.

    Federal procurement (FAR Part 25) requires TAA compliance, so
    classifying equipment as federally procurable implies TAA
    compliance. `commercial` fab_tier doesn't claim either way →
    False (= unverified).
    """
    return spec.get("fab_tier") in {"federal_civilian", "dod_eligible"}


def _plate_spec_to_custom_line(
    plate_id: str,
    plate_spec: dict,
    plate_step_url: str,
    qty: int,
    deployment_context: str = "commercial",
) -> BomLineItem:
    """Map plate spec.yaml + URL → custom_fabrication BOM line item."""
    revision = "001"  # v1 — pull from plate_spec when versioning lands
    pn = f"ARC-PLT-{plate_id}-{revision}"
    if deployment_context != "commercial":
        pn += "-D"

    ctx = plate_spec.get("deployment_contexts", {}).get(deployment_context, {})
    return BomLineItem(
        part_number=pn,
        vendor="ARCNODE (custom fab)",
        description=plate_spec.get("description", f"Interface Plate, {plate_id}"),
        qty=qty,
        procurement_path=ProcurementPath.CUSTOM_FABRICATION,
        material=ctx.get("material"),
        finish=ctx.get("finish"),
        drawing_ref=f"{pn}.dxf",
        drawing_url=plate_step_url.replace(".step", ".dxf"),
    )


class BomGeneratorService:
    """Generates a deployment BOM from manifest + per-assembly bom.yaml."""

    def __init__(self, manifest_client: ManifestClient) -> None:
        self._client = manifest_client

    def generate(
        self,
        *,
        deployment_id: UUID,
        profile: str,
        compute_container_qty: int = 1,
        grid_container_qty: int = 1,
        deployment_context: str = "commercial",
    ) -> Bom:
        """Build a Bom for the given deployment.

        Args:
            deployment_id: UUID of the deployment job.
            profile: Profile name (e.g. "commercial_ac"). Must exist in manifest.
            compute_container_qty: Number of compute containers.
            grid_container_qty: Number of grid containers (0 if no_bess).
            deployment_context: Drives plate variant material/finish.

        Returns:
            Populated Bom ready for serialization.
        """
        manifest = self._client.fetch_manifest()
        if profile not in manifest.profiles:
            raise ValueError(
                f"profile {profile!r} not in manifest (available: {sorted(manifest.profiles)})"
            )
        prof = manifest.profiles[profile]

        line_items: list[BomLineItem] = []
        line_items.extend(self._compute_lines(manifest, prof, compute_container_qty))
        if prof.grid_container is not None and grid_container_qty > 0:
            line_items.extend(self._grid_lines(manifest, prof, grid_container_qty))
        line_items.extend(self._plate_lines(manifest, prof, deployment_context))

        return Bom(
            deployment_id=deployment_id,
            profile=profile,
            manifest_version=manifest.version,
            generated_at=datetime.now(UTC),
            compute_container_qty=compute_container_qty,
            grid_container_qty=grid_container_qty,
            line_items=line_items,
        )

    def _compute_lines(
        self, manifest: Manifest, prof: ProfileAssemblies, container_qty: int
    ) -> list[BomLineItem]:
        cc_variant = manifest.assemblies.get("compute_container", {}).get(
            prof.compute_container
        )
        if cc_variant is None:
            logger.warning(
                f"compute_container variant {prof.compute_container} missing"
            )
            return []
        bom_yaml = self._client.fetch_bom_yaml(cc_variant.bom)
        return self._parts_to_lines(manifest, bom_yaml.get("parts", []), container_qty)

    def _grid_lines(
        self, manifest: Manifest, prof: ProfileAssemblies, container_qty: int
    ) -> list[BomLineItem]:
        if prof.grid_container is None:
            return []
        gc_variant = manifest.assemblies.get("grid_container", {}).get(
            prof.grid_container
        )
        if gc_variant is None:
            logger.warning(
                f"grid_container variant {prof.grid_container} missing — skipping (step 6.1)"
            )
            return []
        bom_yaml = self._client.fetch_bom_yaml(gc_variant.bom)
        return self._parts_to_lines(manifest, bom_yaml.get("parts", []), container_qty)

    def _parts_to_lines(
        self, manifest: Manifest, parts: list[dict], container_qty: int
    ) -> list[BomLineItem]:
        lines: list[BomLineItem] = []
        for part in parts:
            equipment_id = part["equipment_id"]
            per_container_qty = part["qty"]
            spec_url = manifest.specs.get(equipment_id)
            if spec_url is None:
                logger.warning(f"spec URL missing for {equipment_id}")
                continue
            spec = self._client.fetch_spec(spec_url)
            lines.append(_spec_to_catalog_line(spec, per_container_qty * container_qty))
        return lines

    def _plate_lines(
        self,
        manifest: Manifest,
        prof: ProfileAssemblies,
        deployment_context: str,
    ) -> list[BomLineItem]:
        lines: list[BomLineItem] = []
        for plate_id in prof.interface_plates:
            urls = manifest.plates.get(plate_id)
            if urls is None:
                logger.warning(f"plate {plate_id} not in manifest")
                continue
            plate_spec = self._client.fetch_spec(urls.spec)
            lines.append(
                _plate_spec_to_custom_line(
                    plate_id=plate_id,
                    plate_spec=plate_spec,
                    plate_step_url=urls.step,
                    qty=1,
                    deployment_context=deployment_context,
                )
            )
        return lines


def serialize_bom(bom: Bom) -> bytes:
    """Serialize a Bom to JSON bytes for S3 upload."""
    return json.dumps(bom.model_dump(mode="json"), indent=2).encode("utf-8")


# Column order matches BomLineItem field order; metadata cols last so a
# consumer eyeballing the sheet sees procurement essentials first.
_XLSX_COLUMNS: tuple[tuple[str, str], ...] = (
    ("part_number", "Part Number"),
    ("vendor", "Vendor"),
    ("description", "Description"),
    ("qty", "Qty"),
    ("procurement_path", "Procurement Path"),
    ("unit_cost_usd", "Unit Cost (USD)"),
    ("lead_time_weeks", "Lead Time (weeks)"),
    ("datasheet_url", "Datasheet"),
    ("install_video_url", "Install Video"),
    ("ndaa_compliant", "NDAA"),
    ("taa_compliant", "TAA"),
    # Track-B derived columns from `offers`: cheapest live price + source.
    # Empty when enrichment didn't run for this row (no distributor returned
    # a non-error offer). Full per-distributor breakdown lives in the json.
    ("__live_cheapest", "Live Cheapest (USD)"),
    ("__live_source", "Live Source"),
    ("price_change_pct_7d", "Δ vs 7d ago (%)"),
    ("material", "Material"),
    ("finish", "Finish"),
    ("drawing_ref", "Drawing Ref"),
    ("drawing_url", "Drawing URL"),
)


def _cheapest_offer(offers: list) -> tuple[float | None, str | None]:  # type: ignore[type-arg]
    """Lowest non-error offer's (unit_cost_usd, distributor). None when no priced offers."""
    priced = [o for o in offers if o.error is None and o.unit_cost_usd is not None]
    if not priced:
        return None, None
    best = min(priced, key=lambda o: o.unit_cost_usd)
    return best.unit_cost_usd, best.distributor


def serialize_bom_xlsx(bom: Bom) -> bytes:
    """Serialize a Bom to xlsx bytes for S3 upload.

    One header row + one data row per BomLineItem. Header row is bold.
    Empty cells for fields that don't apply to a given procurement_path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    bold = Font(bold=True)
    for col_idx, (_field, label) in enumerate(_XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = bold

    for row_idx, item in enumerate(bom.line_items, start=2):
        cheapest_price, cheapest_source = _cheapest_offer(item.offers)
        for col_idx, (field, _label) in enumerate(_XLSX_COLUMNS, start=1):
            if field == "__live_cheapest":
                value = cheapest_price
            elif field == "__live_source":
                value = cheapest_source
            else:
                value = getattr(item, field)
            # Reason: openpyxl writes StrEnum as the enum object, not its value.
            if hasattr(value, "value"):
                value = value.value
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
