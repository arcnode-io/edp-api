"""InstallSequenceService — narrative Method of Procedure (MOP) PDF.

Walks the DTM, expands per-device install_tasks, groups by BICSI
commissioning level into phase sections (L1 = setup/energization,
L2 = wiring, L3 = functional, L4 = integrated, L5 = performance).
Each section is a numbered checklist with sign-off boxes — the
standard install deliverable handed to site contractors.

PDF only — this is a project-management artifact, not an engineering
drawing. Rendered with reportlab Platypus (no system deps).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from src.shared.schemas.dtm import Dtm
from src.shared.schemas.install_task import CxLevel

_PHASE_TITLE: dict[CxLevel, str] = {
    CxLevel.L1: "PHASE 1 — Site Setup & Energization (L1)",
    CxLevel.L2: "PHASE 2 — Point-to-Point Wiring (L2)",
    CxLevel.L3: "PHASE 3 — Functional Commissioning (L3)",
    CxLevel.L4: "PHASE 4 — Integrated Systems Test (L4)",
    CxLevel.L5: "PHASE 5 — Performance & Sign-Off (L5)",
}


class InstallSequenceOutputs(BaseModel):
    """Bundle of rendered formats from one MOP build."""

    pdf: bytes


class MopRow(BaseModel):
    """One step in the MOP — flat row for the per-phase table."""

    phase: CxLevel
    device_id: str
    task_name: str
    est_minutes: int
    crew_role: str


def build_mop_rows(dtm: Dtm) -> list[MopRow]:
    """Walk DTM → per-device install_tasks → flat rows sorted by (phase, device, name)."""
    rows: list[MopRow] = []
    for device_id in sorted(dtm.devices):
        device = dtm.devices[device_id]
        template = dtm.templates_used[device.template]
        rows.extend(
            MopRow(
                phase=task.cx_level,
                device_id=device_id,
                task_name=task.name,
                est_minutes=task.est_minutes,
                crew_role=task.crew_role,
            )
            for task in template.install_tasks
        )
    rows.sort(key=lambda r: (r.phase.value, r.device_id, r.task_name))
    return rows


class InstallSequenceService:
    """Builds the install sequence MOP PDF from a DTM."""

    def generate(self, dtm: Dtm, profile: str = "") -> InstallSequenceOutputs:
        """Build MOP rows from DTM, render PDF, return as Outputs bundle."""
        rows = build_mop_rows(dtm)
        pdf = _render_mop_pdf(
            rows, deployment_uuid=str(dtm.deployment_uuid), profile=profile
        )
        return InstallSequenceOutputs(pdf=pdf)


def _render_mop_pdf(rows: list[MopRow], *, deployment_uuid: str, profile: str) -> bytes:
    """Reportlab Platypus: title page + one section per phase with step table."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=LETTER,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        title="ARCNODE Installation Method of Procedure",
    )
    styles = getSampleStyleSheet()
    story: list = []
    story.append(
        Paragraph("ARCNODE INSTALLATION — METHOD OF PROCEDURE", styles["Title"])
    )
    story.append(Paragraph(f"Deployment: {deployment_uuid}", styles["Normal"]))
    if profile:
        story.append(Paragraph(f"Profile: {profile}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))
    story.append(
        Paragraph(
            "Phases follow BICSI commissioning convention. Complete each phase "
            "and obtain sign-off before starting the next. Reference: BOM "
            "(bom.xlsx), Cable & Hose Schedule (cable_hose_schedule.xlsx), "
            "SLD (sld.pdf), P&ID (pid_cooling.pdf), Comms Diagram (comms.pdf).",
            styles["Italic"],
        )
    )
    for phase in sorted({r.phase for r in rows}):
        story.append(PageBreak())
        story.append(Paragraph(_PHASE_TITLE[phase], styles["Heading1"]))
        story.append(Spacer(1, 0.1 * inch))
        story.extend(_phase_table(phase, rows))
    doc.build(story)
    return buf.getvalue()


_XLSX_COLUMNS: tuple[tuple[str, str], ...] = (
    ("phase", "Phase"),
    ("device_id", "Device"),
    ("task_name", "Task"),
    ("est_minutes", "Est (min)"),
    ("crew_role", "Crew"),
)


def serialize_install_sequence_xlsx(dtm: Dtm) -> bytes:
    """One Steps sheet — phase / device / task / minutes / crew + Sign-off column."""
    rows = build_mop_rows(dtm)
    wb = Workbook()
    ws = wb.active
    ws.title = "Steps"
    bold = Font(bold=True)
    headers = [*[label for _, label in _XLSX_COLUMNS], "Sign-off"]
    for col_idx, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=label).font = bold
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (field, _label) in enumerate(_XLSX_COLUMNS, start=1):
            value = getattr(row, field)
            if hasattr(value, "value"):
                value = value.value
            ws.cell(row=row_idx, column=col_idx, value=value)
        # Sign-off column — empty cell for installer to mark
        ws.cell(row=row_idx, column=len(headers), value="")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _phase_table(phase: CxLevel, all_rows: list[MopRow]) -> list:
    """Reportlab Table for one phase's steps, plus a phase-total footer."""
    phase_rows = [r for r in all_rows if r.phase == phase]
    header = ["#", "Device", "Task", "Est (min)", "Crew", "Sign-off"]
    data: list[list[str]] = [header]
    for i, r in enumerate(phase_rows, start=1):
        data.append(
            [str(i), r.device_id, r.task_name, str(r.est_minutes), r.crew_role, "☐"]
        )
    total = sum(r.est_minutes for r in phase_rows)
    data.append(["", "", "Phase total", str(total), "", ""])
    table = Table(
        data,
        colWidths=[
            0.4 * inch,
            1.4 * inch,
            2.4 * inch,
            0.9 * inch,
            1.0 * inch,
            0.8 * inch,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#22324c")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("ALIGN", (5, 0), (5, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2f7")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )
    return [table]
