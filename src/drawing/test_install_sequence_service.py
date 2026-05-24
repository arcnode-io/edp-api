"""InstallSequenceService unit tests — narrative MOP (Method of Procedure) PDF.

Standard installation deliverable: PDF organized by commissioning-level
phases (L1→L5), with numbered steps per device task, est minutes,
crew role, and sign-off boxes. Built with reportlab.
"""

import pytest

from src.drawing.conftest import make_device, make_dtm
from src.drawing.install_sequence_service import (
    InstallSequenceOutputs,
    InstallSequenceService,
    build_mop_rows,
    serialize_install_sequence_xlsx,
)
from src.shared.schemas.dtm import Dtm
from src.shared.schemas.install_task import CxLevel, InstallTask
from src.shared.schemas.measurement import Measurement, Publisher
from src.shared.schemas.template import DeviceTemplate, TemplateKind


def _module_tpl(slug: str) -> DeviceTemplate:
    return DeviceTemplate(
        template=slug,
        kind=TemplateKind.MODULE,
        description=slug,
        measurements={
            "total": Measurement(
                unit="watts", type="float", publisher=Publisher.LINE_CONTROLLER
            )
        },
        install_tasks=[
            InstallTask(
                name="container_set",
                est_minutes=240,
                crew_role="general",
                cx_level=CxLevel.L1,
            ),
        ],
    )


def _pdu_tpl() -> DeviceTemplate:
    from src.shared.schemas.template_protocols import ModbusBinding

    return DeviceTemplate(
        template="pdu",
        kind=TemplateKind.LEAF,
        equipment_id="CMP-PDU-001",
        vendor="Server Tech",
        model="PRO3X",
        description="pdu test fixture",
        measurements={
            "v": Measurement(
                unit="volts",
                type="float",
                binding=ModbusBinding(
                    protocol="modbus_tcp", function_code=4, address=100
                ),
            )
        },
        install_tasks=[
            InstallTask(
                name="mount_in_rack",
                est_minutes=20,
                crew_role="electrician",
                cx_level=CxLevel.L1,
            ),
            InstallTask(
                name="wire_input_feed",
                depends_on=["mount_in_rack"],
                est_minutes=30,
                crew_role="electrician",
                cx_level=CxLevel.L2,
            ),
            InstallTask(
                name="snmp_commission",
                depends_on=["wire_input_feed"],
                est_minutes=15,
                crew_role="it",
                cx_level=CxLevel.L3,
            ),
        ],
    )


@pytest.fixture
def small_dtm() -> Dtm:
    """1 compute module + 2 PDUs as children — minimal MOP fixture."""
    return make_dtm(
        devices={
            "compute_module_1": make_device(
                "compute_module_1", template="compute_module"
            ).model_copy(update={"connection": None, "parent": None}),
            "pdu_a": make_device("pdu_a", template="pdu").model_copy(
                update={"parent": "compute_module_1"}
            ),
            "pdu_b": make_device("pdu_b", template="pdu").model_copy(
                update={"parent": "compute_module_1"}
            ),
        },
        templates={
            "compute_module": _module_tpl("compute_module"),
            "pdu": _pdu_tpl(),
        },
    )


def test_generate_returns_pdf_bytes(small_dtm: Dtm) -> None:
    """generate() returns a PDF byte stream."""
    # Act
    actual = InstallSequenceService().generate(small_dtm)

    # Assert
    assert isinstance(actual, InstallSequenceOutputs)
    assert actual.pdf.startswith(b"%PDF-")


def test_mop_rows_grouped_by_cx_level_phase(small_dtm: Dtm) -> None:
    """Rows group by cx_level, sorted L1→L5; per-device sub-order is deterministic.

    Fixture has 3 L1 tasks (compute_module container_set + 2 PDU mounts),
    2 L2 tasks (PDU wire_input_feed x 2), 2 L3 tasks (PDU snmp_commission x 2).
    """
    # Act
    rows = build_mop_rows(small_dtm)

    # Assert
    phases = [r.phase for r in rows]
    assert phases.count(CxLevel.L1) == 3
    assert phases.count(CxLevel.L2) == 2
    assert phases.count(CxLevel.L3) == 2
    # L1 before L2 before L3 (sorted by phase enum order in output).
    l1_end = max(i for i, r in enumerate(rows) if r.phase == CxLevel.L1)
    l2_start = min(i for i, r in enumerate(rows) if r.phase == CxLevel.L2)
    assert l1_end < l2_start


def test_xlsx_starts_with_zip_magic_and_has_steps_sheet(small_dtm: Dtm) -> None:
    """xlsx artifact is a real openpyxl workbook with a Steps sheet, all rows present."""
    from io import BytesIO

    from openpyxl import load_workbook

    # Act
    xlsx_bytes = serialize_install_sequence_xlsx(small_dtm)

    # Assert — zip magic + sheet name + row count = 1 header + N data rows.
    assert xlsx_bytes.startswith(b"PK")
    wb = load_workbook(BytesIO(xlsx_bytes))
    assert "Steps" in wb.sheetnames
    ws = wb["Steps"]
    # Fixture has 3 + 2 + 2 = 7 rows + header = 8 total.
    assert ws.max_row == 8


def test_pdf_has_one_page_per_phase_plus_cover(small_dtm: Dtm) -> None:
    """PDF page count = 1 cover + N phases (fixture spans L1+L2+L3 = 3 phases)."""
    import re

    # Act
    pdf = InstallSequenceService().generate(small_dtm).pdf

    # Assert
    page_count = len(re.findall(rb"/Type\s*/Page(?!s)", pdf))
    assert page_count == 4  # 1 cover + 3 phase sections
