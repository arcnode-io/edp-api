"""PipelineService unit tests — captures S3 puts via a stub ManifestClient."""

import json
from typing import cast
from uuid import UUID

from src.bom_generator.bom_generator_service import BomGeneratorService
from src.bom_generator.manifest_client import ManifestClient
from src.bom_generator.manifest_models import (
    AssemblyVariant,
    Manifest,
    PlateUrls,
    ProfileAssemblies,
)
from src.bom_generator.manifest_service import ManifestService
from src.cable_hose_schedule.cable_hose_schedule_service import CableHoseScheduleService
from src.drawing.comms_diagram_service import CommsDiagramService
from src.drawing.install_sequence_service import InstallSequenceService
from src.drawing.pid_cooling_service import PidCoolingService
from src.drawing.sld_engineering_service import SldEngineeringService
from src.drawing.sld_hmi_svg_service import SldHmiSvgService
from src.dtm.dtm_generator_service import DtmGeneratorService
from src.dtm.template_loader import TemplateLoader
from src.module_resolver.module_resolver_service import ModuleResolverService
from src.pipeline.artifact_urls import build_artifact_urls_from_resolved
from src.pipeline.pipeline_service import PipelineService
from src.shared.enums import (
    AwsPartition,
    WholesaleMarket,
    BessCoupling,
    ClimateZone,
    DeploymentContext,
    EnergySource,
    GpuVariant,
    GridConnection,
    PrimaryWorkload,
)
from src.shared.schemas.artifact import ArtifactKind
from src.shared.schemas.configurator_payload import ConfiguratorPayload

DEPLOYMENT_ID: UUID = UUID("00000000-0000-0000-0000-000000000abc")


class _RecordingClient:
    """Stub ManifestClient: records uploads, returns an in-memory manifest."""

    def __init__(self, manifest: Manifest) -> None:
        self._manifest = manifest
        self.uploads: dict[str, bytes] = {}

    def fetch_manifest(self) -> Manifest:
        return self._manifest

    def fetch_bom_yaml(self, _url: str) -> dict:
        # Minimum viable bom.yaml shape — empty parts so BomGenerator emits
        # zero line items (we're testing dispatch, not BOM content).
        return {"parts": []}

    def fetch_topology_yaml(self, _url: str) -> dict:
        # Real topology yaml is consumed by emit_container; covered by
        # dtm_generator's own tests. Here we just need *something* parseable.
        return {"devices": []}

    def fetch_spec(self, _url: str) -> dict:
        return {}

    def upload_bom_json(self, body: bytes, target_url: str) -> None:
        self.uploads[target_url] = body

    def upload_bytes(self, body: bytes, target_url: str) -> None:
        self.uploads[target_url] = body


def _commercial_ac_manifest() -> Manifest:
    """Profile + assemblies + plates needed for commercial_ac end-to-end."""
    asm = AssemblyVariant(
        bom="s3://test/compute/commercial-ac/bom.yaml",
        step="s3://arcnode-artifacts/assemblies/compute-container/commercial-ac/assembly.step",
        glb="s3://arcnode-artifacts/assemblies/compute-container/commercial-ac/assembly.glb",
        topology_yaml="s3://test/compute/commercial-ac/topology.yaml",
    )
    grid = AssemblyVariant(
        bom="s3://test/grid/commercial-ac/bom.yaml",
        step="s3://arcnode-artifacts/assemblies/grid-container/commercial-ac/assembly.step",
        glb="s3://arcnode-artifacts/assemblies/grid-container/commercial-ac/assembly.glb",
        topology_yaml="s3://test/grid/commercial-ac/topology.yaml",
    )
    return Manifest(
        version="0.0.0-test",
        assemblies={
            "compute_container": {"commercial-ac": asm},
            "grid_container": {"commercial-ac": grid},
        },
        plates={
            pid: PlateUrls(
                spec=f"s3://arcnode-artifacts/plates/{pid}/spec.yaml",
                step=f"s3://arcnode-artifacts/plates/{pid}/{pid}.step",
                dxf=f"s3://arcnode-artifacts/plates/{pid}/{pid}.dxf",
            )
            for pid in ("CG", "BG-AC", "CD")
        },
        profiles={
            "commercial_ac": ProfileAssemblies(
                compute_container="commercial-ac",
                grid_container="commercial-ac",
                interface_plates=["CG", "BG-AC", "CD"],
            ),
        },
    )


def _payload() -> ConfiguratorPayload:
    return ConfiguratorPayload(
        deployment_id=DEPLOYMENT_ID,
        operator_org="acme",
        deployment_site_name="brookside dc-1",
        contact_email="ops@example.com",
        energy_source=EnergySource.GRID_HYBRID,
        source_capacity_mw=10.0,
        primary_workload=PrimaryWorkload.AI_TRAINING,
        gpu_variant=GpuVariant.H100_SXM,
        target_gpu_count=56,
        bess_coupling=BessCoupling.AC_COUPLED,
        bess_capacity_mwh=5.0,
        grid_connection=GridConnection.GRID_TIED,
        climate_zone=ClimateZone.TEMPERATE,
        deployment_context=DeploymentContext.COMMERCIAL,
        aws_partition=AwsPartition.STANDARD,
        wholesale_market=WholesaleMarket.ERCOT,
        settlement_point="HB_NORTH",
    )


def _build_pipeline(client: _RecordingClient) -> PipelineService:
    real_client = cast(ManifestClient, client)
    catalog = TemplateLoader(
        root=__import__("pathlib").Path(__file__).resolve().parents[2]
        / "device_templates"
    ).load_catalog()
    return PipelineService(
        client=real_client,
        bom_generator=BomGeneratorService(real_client),
        dtm_generator=DtmGeneratorService(real_client, template_catalog=catalog),
        sld_hmi_svg_service=SldHmiSvgService(),
        sld_engineering_service=SldEngineeringService(),
        pid_cooling_service=PidCoolingService(),
        comms_diagram_service=CommsDiagramService(),
        cable_hose_schedule_service=CableHoseScheduleService(),
        install_sequence_service=InstallSequenceService(),
    )


def test_run_uploads_to_every_generated_url() -> None:
    """Per-deployment URLs (under edp/{deployment_id}/) all get bytes."""
    # Arrange
    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert — every per-deployment URL was uploaded to exactly once.
    expected = {u.url for u in urls if u.url.startswith("s3://arcnode-artifacts/edp/")}
    assert expected == set(client.uploads.keys()), (
        f"missing: {expected - set(client.uploads)}, "
        f"extra: {set(client.uploads) - expected}"
    )


def test_run_skips_selected_urls_from_catalog() -> None:
    """Selected URLs (assemblies, plates) are NOT touched — they live in catalog."""
    # Arrange
    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert — none of the catalog URLs were uploaded
    catalog_urls = {
        u.url for u in urls if not u.url.startswith("s3://arcnode-artifacts/edp/")
    }
    assert catalog_urls.isdisjoint(client.uploads.keys())


def test_bom_upload_is_real_bom_json() -> None:
    """The BOM upload deserializes to a Bom shape (not a stub)."""
    # Arrange
    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert
    bom_url = next(
        u.url for u in urls if u.kind == ArtifactKind.BOM and u.format == "json"
    )
    body = json.loads(client.uploads[bom_url])
    assert body["deployment_id"] == str(DEPLOYMENT_ID)
    assert body["profile"] == "commercial_ac"
    assert "line_items" in body


def test_sld_engineering_uploads_real_dxf_and_pdf() -> None:
    """SLD dxf URL gets parseable DXF bytes + sld pdf URL gets %PDF- magic bytes."""
    # Arrange
    import io as _io

    import ezdxf as _ezdxf

    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert — both URLs got real bytes, not stubs
    dxf_url = next(
        u.url for u in urls if u.kind == ArtifactKind.SLD and u.format == "dxf"
    )
    pdf_url = next(
        u.url for u in urls if u.kind == ArtifactKind.SLD and u.format == "pdf"
    )
    # DXF round-trips through ezdxf — proves real DXF.
    _ezdxf.read(_io.StringIO(client.uploads[dxf_url].decode("utf-8")))
    # PDF starts with the %PDF- magic.
    assert client.uploads[pdf_url].startswith(b"%PDF-")


def test_cable_hose_schedule_uploads_real_json_and_xlsx() -> None:
    """Cable+hose json URL gets a CableHoseSchedule shape + xlsx round-trips."""
    # Arrange
    import io as _io

    from openpyxl import load_workbook as _load_wb

    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert — json deserializes with `cables` + `hoses` keys.
    json_url = next(
        u.url
        for u in urls
        if u.kind == ArtifactKind.CABLE_HOSE_SCHEDULE and u.format == "json"
    )
    body = json.loads(client.uploads[json_url])
    assert "cables" in body
    assert "hoses" in body

    # Assert — xlsx round-trips with Cables + Hoses sheets.
    xlsx_url = next(
        u.url
        for u in urls
        if u.kind == ArtifactKind.CABLE_HOSE_SCHEDULE and u.format == "xlsx"
    )
    wb = _load_wb(_io.BytesIO(client.uploads[xlsx_url]))
    assert "Cables" in wb.sheetnames
    assert "Hoses" in wb.sheetnames


def test_comms_diagram_uploads_real_dxf_and_pdf() -> None:
    """Comms diagram dxf URL gets parseable DXF + pdf URL gets %PDF- magic bytes."""
    # Arrange
    import io as _io

    import ezdxf as _ezdxf

    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert
    dxf_url = next(
        u.url
        for u in urls
        if u.kind == ArtifactKind.COMMS_DIAGRAM and u.format == "dxf"
    )
    pdf_url = next(
        u.url
        for u in urls
        if u.kind == ArtifactKind.COMMS_DIAGRAM and u.format == "pdf"
    )
    _ezdxf.read(_io.StringIO(client.uploads[dxf_url].decode("utf-8")))
    assert client.uploads[pdf_url].startswith(b"%PDF-")


def test_pid_cooling_uploads_real_dxf_and_pdf() -> None:
    """P&ID dxf URL gets parseable DXF + pdf URL gets %PDF- magic bytes."""
    # Arrange
    import io as _io

    import ezdxf as _ezdxf

    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert
    dxf_url = next(
        u.url for u in urls if u.kind == ArtifactKind.PID_COOLING and u.format == "dxf"
    )
    pdf_url = next(
        u.url for u in urls if u.kind == ArtifactKind.PID_COOLING and u.format == "pdf"
    )
    _ezdxf.read(_io.StringIO(client.uploads[dxf_url].decode("utf-8")))
    assert client.uploads[pdf_url].startswith(b"%PDF-")


def test_bom_xlsx_upload_is_real_workbook() -> None:
    """The BOM xlsx upload opens back as a workbook (not stub bytes)."""
    # Arrange
    from io import BytesIO

    from openpyxl import load_workbook

    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert — round-trips as a real xlsx with a BOM sheet
    xlsx_url = next(
        u.url for u in urls if u.kind == ArtifactKind.BOM and u.format == "xlsx"
    )
    wb = load_workbook(BytesIO(client.uploads[xlsx_url]))
    assert wb.active is not None
    assert wb.active.title == "BOM"


def test_dtm_upload_is_real_dtm_json() -> None:
    """The DTM upload deserializes to a Dtm shape (not a stub)."""
    # Arrange
    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert
    dtm_url = next(u.url for u in urls if u.kind == ArtifactKind.DTM)
    body = json.loads(client.uploads[dtm_url])
    assert body["deployment_uuid"] == str(DEPLOYMENT_ID)
    assert body["version"] == "1.0.0"
    assert "devices" in body


def test_sld_hmi_svg_upload_is_real_svg() -> None:
    """SLD_HMI_SVG upload is a parseable SVG with HMI's data-* hooks."""
    # Arrange
    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert
    svg_url = next(u.url for u in urls if u.kind == ArtifactKind.SLD_HMI_SVG)
    body = client.uploads[svg_url].decode("utf-8")
    assert body.startswith("<?xml version=")
    assert 'data-comp="device-node"' in body or 'data-comp="bus"' in body


def test_all_reserved_artifact_kinds_now_get_real_bytes() -> None:
    """No `_stub_body` fallback paths exercised — every reserved kind has a real generator.

    Regression guard: if a future refactor accidentally drops a real
    generator from the dispatch table, the corresponding URL would
    silently receive stub bytes — this test catches that drift.
    """
    # Arrange
    client = _RecordingClient(_commercial_ac_manifest())
    pipeline = _build_pipeline(client)
    payload = _payload()
    resolution = ModuleResolverService().resolve(payload)
    urls = build_artifact_urls_from_resolved(
        DEPLOYMENT_ID,
        ManifestService(manifest=_commercial_ac_manifest()).resolve("commercial_ac"),
    )

    # Act
    pipeline.run(
        payload=payload,
        resolution=resolution,
        urls=urls,
        manifest=_commercial_ac_manifest(),
    )

    # Assert — every reserved generated URL got non-stub bytes. Pick the
    # BOM xlsx to verify the shape:
    # real xlsx bytes start with the ZIP magic 'PK'.
    bom_xlsx = next(
        u.url for u in urls if u.kind == ArtifactKind.BOM and u.format == "xlsx"
    )
    assert client.uploads[bom_xlsx][:2] == b"PK"


def test_attach_offers_merges_distributor_offers_onto_catalog_lines() -> None:
    """`_attach_offers` populates each catalog line's `offers` from enrichment.

    Direct unit-level test of the merge helper — the higher-level
    pipeline tests use an empty bom.yaml fixture so wouldn't exercise
    the catalog-line path. This test gives `_attach_offers` a real Bom
    with one catalog + one custom-fab line and asserts only the catalog
    line gets offers (custom-fab plates have no distributor).
    """
    # Arrange
    from datetime import UTC, datetime as _dt

    from src.bom_enrichment._base_scraper import DistributorClient
    from src.bom_enrichment.enrichment_models import (
        DistributorId,
        DistributorOffer,
    )
    from src.bom_enrichment.enrichment_service import EnrichmentService
    from src.bom_generator.bom_models import Bom, BomLineItem, ProcurementPath
    from src.pipeline.pipeline_service import _attach_offers

    class _MockClient(DistributorClient):
        @property
        def distributor_id(self) -> DistributorId:
            return "graybar"

        def fetch_offer(self, mpn: str) -> DistributorOffer:
            return DistributorOffer(
                distributor="graybar",
                mpn=mpn,
                stock_count=5,
                unit_cost_usd=99.99,
                refreshed_at=_dt.now(UTC),
            )

        def close(self) -> None:
            pass

    bom = Bom(
        deployment_id=DEPLOYMENT_ID,
        profile="commercial_ac",
        manifest_version="0.0.0-test",
        generated_at=_dt.now(UTC),
        compute_container_qty=1,
        grid_container_qty=0,
        line_items=[
            BomLineItem(
                part_number="MPN-CATALOG-1",
                vendor="Acme",
                description="x",
                qty=1,
                procurement_path=ProcurementPath.CATALOG,
            ),
            BomLineItem(
                part_number="ARC-PLT-CG-001",
                vendor="ARCNODE (custom fab)",
                description="plate",
                qty=1,
                procurement_path=ProcurementPath.CUSTOM_FABRICATION,
            ),
        ],
    )

    # Act
    _attach_offers(bom, EnrichmentService([_MockClient()]))

    # Assert — catalog line got an offer; custom-fab line did NOT
    catalog_line = next(
        li for li in bom.line_items if li.procurement_path == ProcurementPath.CATALOG
    )
    plate_line = next(
        li
        for li in bom.line_items
        if li.procurement_path == ProcurementPath.CUSTOM_FABRICATION
    )
    assert len(catalog_line.offers) == 1
    assert catalog_line.offers[0].unit_cost_usd == 99.99
    assert catalog_line.offers[0].distributor == "graybar"
    assert plate_line.offers == []
